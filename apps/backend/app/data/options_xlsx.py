from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]

from app.schema.options_models import OptionsRecord


DATA_DIR = Path(__file__).resolve().parent
XLSX_PATH = DATA_DIR / "bonds.xlsx"
SHEET_NAME = "options"

HEADERS = ["year", "amount"]


def _ensure_workbook_exists() -> None:
    """Ensure the bonds.xlsx workbook and options sheet exist on disk."""

    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
    else:
        wb = Workbook()
        # Name the default sheet "bonds" to stay consistent with bonds_xlsx
        ws = wb.active
        ws.title = "bonds"
        wb.save(XLSX_PATH)

    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        wb.save(XLSX_PATH)


def _get_sheet() -> Worksheet:
    _ensure_workbook_exists()
    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]  # type: ignore[return-value]
    return wb.active  # type: ignore[return-value]


def load_options() -> List[OptionsRecord]:
    """Load options income records from the options sheet."""

    ws = _get_sheet()

    header_row = [cell.value for cell in ws[1]]
    indices = {name: header_row.index(name) for name in HEADERS if name in header_row}

    records: List[OptionsRecord] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        def col(name: str):
            idx = indices.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        year_raw = col("year")
        amount_raw = col("amount")

        if year_raw is None or amount_raw is None:
            continue

        try:
            year = int(year_raw)
            amount = float(amount_raw)
        except Exception:
            continue

        records.append(OptionsRecord(year=year, amount=amount))

    records.sort(key=lambda r: r.year)
    return records


def save_options(records: List[OptionsRecord]) -> None:
    """Persist the given options income records to the options sheet."""

    _ensure_workbook_exists()
    wb = load_workbook(XLSX_PATH)

    if SHEET_NAME in wb.sheetnames:
        wb.remove(wb[SHEET_NAME])

    ws = wb.create_sheet(SHEET_NAME)

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row_idx = 2
    for record in records:
        ws.cell(row=row_idx, column=1, value=record.year)
        ws.cell(row=row_idx, column=2, value=record.amount)
        row_idx += 1

    wb.save(XLSX_PATH)

