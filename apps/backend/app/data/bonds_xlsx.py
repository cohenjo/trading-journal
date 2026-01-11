from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]

from .bonds_types import BondHolding


DATA_DIR = Path(__file__).resolve().parent
XLSX_PATH = DATA_DIR / "bonds.xlsx"
SHEET_NAME = "bonds"

HEADERS = [
    "id",
    "ticker",
    "issuer",
    "currency",
    "face_value",
    "coupon_rate",
    "coupon_frequency",
    "issue_date",
    "maturity_date",
]


def _ensure_workbook_exists(initial_bonds: Iterable[BondHolding]) -> None:
    """Create a minimal workbook on disk if it doesn't exist yet.

    If the file already exists, this is a no-op.
    """

    if XLSX_PATH.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Seed with initial bonds, if any
    row_idx = 2
    for b in initial_bonds:
        ws.cell(row=row_idx, column=1, value=b.id)
        ws.cell(row=row_idx, column=2, value=b.ticker)
        ws.cell(row=row_idx, column=3, value=b.issuer)
        ws.cell(row=row_idx, column=4, value=b.currency)
        ws.cell(row=row_idx, column=5, value=b.face_value)
        ws.cell(row=row_idx, column=6, value=b.coupon_rate)
        ws.cell(row=row_idx, column=7, value=b.coupon_frequency)
        ws.cell(row=row_idx, column=8, value=b.issue_date.isoformat())
        ws.cell(row=row_idx, column=9, value=b.maturity_date.isoformat())
        row_idx += 1

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)


def _get_sheet() -> Worksheet:
    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]  # type: ignore[return-value]
    # Fallback: use active sheet if name changed
    return wb.active  # type: ignore[return-value]


def load_bonds_from_xlsx(initial_bonds: Iterable[BondHolding]) -> List[BondHolding]:
    """Load BondHolding records from the bonds.xlsx file.

    If the file does not exist yet, it is created and seeded with the
    provided initial_bonds, then those are returned.
    """

    _ensure_workbook_exists(initial_bonds)

    ws = _get_sheet()

    # Map header names to column indices
    header_row = [cell.value for cell in ws[1]]
    indices = {name: header_row.index(name) for name in HEADERS if name in header_row}

    bonds: List[BondHolding] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        def col(name: str):
            idx = indices.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        id_val = col("id")
        ticker = col("ticker")
        issuer = col("issuer")
        currency = col("currency") or "USD"
        face_raw = col("face_value")
        coupon_raw = col("coupon_rate")

        def to_float(val) -> float:
            if val in (None, ""):
                return 0.0
            try:
                return float(val)  # type: ignore[arg-type]
            except Exception:
                return 0.0

        face_value = to_float(face_raw)
        coupon_rate = to_float(coupon_raw)
        coupon_frequency = col("coupon_frequency") or "ANNUAL"

        issue_raw = col("issue_date")
        maturity_raw = col("maturity_date")

        def to_date(val) -> date:
            if isinstance(val, date):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, str):
                return date.fromisoformat(val)
            raise ValueError(f"Unsupported date cell value: {val!r}")

        if not id_val or not issuer or not issue_raw or not maturity_raw:
            # Skip incomplete rows
            continue

        issue_date = to_date(issue_raw)
        maturity_date = to_date(maturity_raw)

        bonds.append(
            BondHolding(
                id=str(id_val),
                ticker=str(ticker) if ticker is not None else None,
                issuer=str(issuer),
                currency=str(currency),
                face_value=face_value,
                coupon_rate=coupon_rate,
                coupon_frequency=str(coupon_frequency),
                issue_date=issue_date,
                maturity_date=maturity_date,
            )
        )

    return bonds


def save_bonds_to_xlsx(bonds: Iterable[BondHolding]) -> None:
    """Persist the given list of BondHolding records to bonds.xlsx."""

    _ensure_workbook_exists(initial_bonds=[])

    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.create_sheet(SHEET_NAME)

    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Write header
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row_idx = 2
    for b in bonds:
        ws.cell(row=row_idx, column=1, value=b.id)
        ws.cell(row=row_idx, column=2, value=b.ticker)
        ws.cell(row=row_idx, column=3, value=b.issuer)
        ws.cell(row=row_idx, column=4, value=b.currency)
        ws.cell(row=row_idx, column=5, value=b.face_value)
        ws.cell(row=row_idx, column=6, value=b.coupon_rate)
        ws.cell(row=row_idx, column=7, value=b.coupon_frequency)
        ws.cell(row=row_idx, column=8, value=b.issue_date.isoformat())
        ws.cell(row=row_idx, column=9, value=b.maturity_date.isoformat())
        row_idx += 1

    wb.save(XLSX_PATH)
