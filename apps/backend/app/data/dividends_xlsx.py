from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]

from ..schema.dividend_models import DividendRecord

DATA_DIR = Path(__file__).resolve().parent
XLSX_PATH = DATA_DIR / "bonds.xlsx"
SHEET_NAME = "dividends"

HEADERS = ["year", "amount"]

def _ensure_workbook_exists() -> None:
    """Create a minimal workbook on disk if it doesn't exist yet.
    If the file already exists, ensure the sheet exists.
    """
    if not XLSX_PATH.exists():
        wb = Workbook()
        # Remove default sheet if we are creating a new one, or rename it
        ws = wb.active
        ws.title = "bonds" # Default to bonds as per other file, or just leave it. 
        # The other file handles creation of 'bonds' sheet. 
        # Let's just save it.
        wb.save(XLSX_PATH)

    wb = load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        wb.save(XLSX_PATH)

def _get_sheet() -> Worksheet:
    _ensure_workbook_exists()
    wb = load_workbook(XLSX_PATH)
    return wb[SHEET_NAME]

def load_dividends() -> List[DividendRecord]:
    """Load DividendRecord records from the xlsx file."""
    _ensure_workbook_exists()
    ws = _get_sheet()

    # Map header names to column indices
    header_row = [cell.value for cell in ws[1]]
    indices = {name: header_row.index(name) for name in HEADERS if name in header_row}

    records: List[DividendRecord] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        def col(name: str):
            idx = indices.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        year_val = col("year")
        amount_val = col("amount")

        if year_val is None or amount_val is None:
            continue

        try:
            records.append(DividendRecord(year=int(year_val), amount=float(amount_val)))
        except ValueError:
            continue
            
    # Sort by year
    records.sort(key=lambda x: x.year)
    return records

def save_dividends(records: List[DividendRecord]) -> None:
    """Persist the given list of DividendRecord records."""
    _ensure_workbook_exists()
    wb = load_workbook(XLSX_PATH)
    
    if SHEET_NAME in wb.sheetnames:
        # Remove old sheet to overwrite
        del wb[SHEET_NAME]
    
    ws = wb.create_sheet(SHEET_NAME)

    # Write header
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row_idx = 2
    for r in records:
        ws.cell(row=row_idx, column=1, value=r.year)
        ws.cell(row=row_idx, column=2, value=r.amount)
        row_idx += 1

    wb.save(XLSX_PATH)
