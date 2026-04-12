"""
Tests for XLSX data loaders: bonds, dividends, and options.

Uses temporary workbook files to verify:
  - Round-trip load/save consistency
  - Type conversions (to_float, to_date)
  - Handling of missing/invalid data
  - Correct header mapping

Each test creates its own workbook in the project dir, patching the module-level
XLSX_PATH so we never touch real data files.
"""

import pytest
from datetime import date
from pathlib import Path
from unittest.mock import patch
from openpyxl import Workbook

from app.data.bonds_types import BondHolding
from app.schema.dividend_models import DividendRecord
from app.schema.options_models import OptionsRecord


# ===================================================================
# Fixtures
# ===================================================================

@pytest.fixture
def xlsx_dir(request):
    """Create a temp directory in the project for XLSX test files."""
    test_dir = Path(__file__).parent / "_xlsx_test_data"
    test_dir.mkdir(exist_ok=True)
    yield test_dir
    # Cleanup after test
    import shutil
    shutil.rmtree(test_dir, ignore_errors=True)


# ===================================================================
# Bonds XLSX Loader
# ===================================================================

class TestBondsXlsxLoader:
    def _create_bonds_xlsx(self, filepath: Path, rows: list[list]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "bonds"
        headers = ["id", "ticker", "issuer", "currency", "face_value",
                    "coupon_rate", "coupon_frequency", "issue_date", "maturity_date"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        wb.save(filepath)

    @patch("app.data.bonds_xlsx.XLSX_PATH")
    def test_load_valid_bonds(self, mock_path, xlsx_dir):
        filepath = xlsx_dir / "bonds_valid.xlsx"
        self._create_bonds_xlsx(filepath, [
            ["B1", "GOVT", "US Treasury", "USD", 10000, 0.05, "SEMI_ANNUAL", "2023-01-01", "2033-01-01"],
            ["B2", None, "Israel Gov", "ILS", 5000, 0.03, "ANNUAL", "2022-06-15", "2032-06-15"],
        ])
        mock_path.__fspath__ = lambda self: str(filepath)
        mock_path.exists = lambda: True
        mock_path.__str__ = lambda self: str(filepath)

        from app.data import bonds_xlsx
        original_path = bonds_xlsx.XLSX_PATH
        bonds_xlsx.XLSX_PATH = filepath
        try:
            bonds = bonds_xlsx.load_bonds_from_xlsx(initial_bonds=[])
            assert len(bonds) == 2
            assert bonds[0].id == "B1"
            assert bonds[0].face_value == 10000.0
            assert bonds[0].coupon_rate == 0.05
            assert bonds[0].issue_date == date(2023, 1, 1)
            assert bonds[1].ticker is None or bonds[1].ticker == "None"
        finally:
            bonds_xlsx.XLSX_PATH = original_path

    @patch("app.data.bonds_xlsx.XLSX_PATH")
    def test_load_skips_incomplete_rows(self, mock_path, xlsx_dir):
        filepath = xlsx_dir / "bonds_incomplete.xlsx"
        self._create_bonds_xlsx(filepath, [
            ["B1", "GOVT", "US Treasury", "USD", 10000, 0.05, "ANNUAL", "2023-01-01", "2033-01-01"],
            [None, None, None, None, None, None, None, None, None],  # Empty row
            ["B3", None, None, None, 5000, 0.03, None, None, None],  # Missing issuer + dates
        ])
        from app.data import bonds_xlsx
        original_path = bonds_xlsx.XLSX_PATH
        bonds_xlsx.XLSX_PATH = filepath
        try:
            bonds = bonds_xlsx.load_bonds_from_xlsx(initial_bonds=[])
            assert len(bonds) == 1
            assert bonds[0].id == "B1"
        finally:
            bonds_xlsx.XLSX_PATH = original_path

    @patch("app.data.bonds_xlsx.XLSX_PATH")
    def test_to_float_defaults_for_invalid(self, mock_path, xlsx_dir):
        """face_value/coupon_rate with None or invalid strings default to 0.0."""
        filepath = xlsx_dir / "bonds_badfloat.xlsx"
        self._create_bonds_xlsx(filepath, [
            ["B1", "T", "Issuer", "USD", "not_a_number", None, "ANNUAL", "2023-01-01", "2033-01-01"],
        ])
        from app.data import bonds_xlsx
        original_path = bonds_xlsx.XLSX_PATH
        bonds_xlsx.XLSX_PATH = filepath
        try:
            bonds = bonds_xlsx.load_bonds_from_xlsx(initial_bonds=[])
            assert len(bonds) == 1
            assert bonds[0].face_value == 0.0
            assert bonds[0].coupon_rate == 0.0
        finally:
            bonds_xlsx.XLSX_PATH = original_path

    @patch("app.data.bonds_xlsx.XLSX_PATH")
    def test_round_trip_save_load(self, mock_path, xlsx_dir):
        filepath = xlsx_dir / "bonds_roundtrip.xlsx"
        from app.data import bonds_xlsx
        original_path = bonds_xlsx.XLSX_PATH
        bonds_xlsx.XLSX_PATH = filepath

        try:
            original_bonds = [
                BondHolding(
                    id="RT1", ticker="BOND1", issuer="Test Issuer",
                    currency="USD", face_value=25000.50, coupon_rate=0.045,
                    coupon_frequency="QUARTERLY",
                    issue_date=date(2024, 1, 15), maturity_date=date(2034, 1, 15),
                ),
            ]
            # Save
            bonds_xlsx._ensure_workbook_exists(initial_bonds=[])
            bonds_xlsx.save_bonds_to_xlsx(original_bonds)
            # Load
            loaded = bonds_xlsx.load_bonds_from_xlsx(initial_bonds=[])
            assert len(loaded) == 1
            assert loaded[0].id == "RT1"
            assert loaded[0].face_value == 25000.50
            assert loaded[0].coupon_rate == 0.045
            assert loaded[0].issue_date == date(2024, 1, 15)
        finally:
            bonds_xlsx.XLSX_PATH = original_path


# ===================================================================
# Dividends XLSX Loader
# ===================================================================

class TestDividendsXlsxLoader:
    def _create_dividends_xlsx(self, filepath: Path, rows: list[list]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "bonds"  # Default sheet per dividends_xlsx.py
        ws_div = wb.create_sheet("dividends")
        ws_div.cell(row=1, column=1, value="year")
        ws_div.cell(row=1, column=2, value="amount")
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws_div.cell(row=row_idx, column=col_idx, value=val)
        wb.save(filepath)

    def test_load_valid_dividends(self, xlsx_dir):
        filepath = xlsx_dir / "div_valid.xlsx"
        self._create_dividends_xlsx(filepath, [
            [2021, 1500.0],
            [2022, 1800.0],
            [2023, 2100.0],
        ])
        from app.data import dividends_xlsx
        original_path = dividends_xlsx.XLSX_PATH
        dividends_xlsx.XLSX_PATH = filepath
        try:
            records = dividends_xlsx.load_dividends()
            assert len(records) == 3
            assert records[0].year == 2021
            assert records[0].amount == 1500.0
            assert records[-1].year == 2023
        finally:
            dividends_xlsx.XLSX_PATH = original_path

    def test_skips_rows_with_none_values(self, xlsx_dir):
        filepath = xlsx_dir / "div_nones.xlsx"
        self._create_dividends_xlsx(filepath, [
            [2021, 1500.0],
            [None, 1800.0],   # Missing year
            [2023, None],     # Missing amount
            [2024, 2000.0],
        ])
        from app.data import dividends_xlsx
        original_path = dividends_xlsx.XLSX_PATH
        dividends_xlsx.XLSX_PATH = filepath
        try:
            records = dividends_xlsx.load_dividends()
            assert len(records) == 2
            assert records[0].year == 2021
            assert records[1].year == 2024
        finally:
            dividends_xlsx.XLSX_PATH = original_path

    def test_skips_invalid_type_conversion(self, xlsx_dir):
        filepath = xlsx_dir / "div_badtype.xlsx"
        self._create_dividends_xlsx(filepath, [
            [2021, 1500.0],
            ["not_a_year", 1000.0],  # ValueError on int()
        ])
        from app.data import dividends_xlsx
        original_path = dividends_xlsx.XLSX_PATH
        dividends_xlsx.XLSX_PATH = filepath
        try:
            records = dividends_xlsx.load_dividends()
            assert len(records) == 1
            assert records[0].year == 2021
        finally:
            dividends_xlsx.XLSX_PATH = original_path

    def test_round_trip_save_load(self, xlsx_dir):
        filepath = xlsx_dir / "div_roundtrip.xlsx"
        from app.data import dividends_xlsx
        original_path = dividends_xlsx.XLSX_PATH
        dividends_xlsx.XLSX_PATH = filepath
        try:
            original = [
                DividendRecord(year=2022, amount=1500.50),
                DividendRecord(year=2023, amount=2000.75),
            ]
            dividends_xlsx.save_dividends(original)
            loaded = dividends_xlsx.load_dividends()
            assert len(loaded) == 2
            assert loaded[0].year == 2022
            assert loaded[0].amount == pytest.approx(1500.50)
            assert loaded[1].year == 2023
        finally:
            dividends_xlsx.XLSX_PATH = original_path


# ===================================================================
# Options XLSX Loader
# ===================================================================

class TestOptionsXlsxLoader:
    def _create_options_xlsx(self, filepath: Path, rows: list[list]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "bonds"
        ws_opt = wb.create_sheet("options")
        ws_opt.cell(row=1, column=1, value="year")
        ws_opt.cell(row=1, column=2, value="amount")
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws_opt.cell(row=row_idx, column=col_idx, value=val)
        wb.save(filepath)

    def test_load_valid_options(self, xlsx_dir):
        filepath = xlsx_dir / "opt_valid.xlsx"
        self._create_options_xlsx(filepath, [
            [2021, 3000.0],
            [2022, 4500.0],
            [2023, 6000.0],
        ])
        from app.data import options_xlsx
        original_path = options_xlsx.XLSX_PATH
        options_xlsx.XLSX_PATH = filepath
        try:
            records = options_xlsx.load_options()
            assert len(records) == 3
            assert records[0].year == 2021
            assert records[-1].amount == 6000.0
        finally:
            options_xlsx.XLSX_PATH = original_path

    def test_skips_invalid_rows(self, xlsx_dir):
        filepath = xlsx_dir / "opt_invalid.xlsx"
        self._create_options_xlsx(filepath, [
            [2021, 3000.0],
            [None, None],         # All None
            ["bad", "data"],      # Can't convert
            [2023, 5000.0],
        ])
        from app.data import options_xlsx
        original_path = options_xlsx.XLSX_PATH
        options_xlsx.XLSX_PATH = filepath
        try:
            records = options_xlsx.load_options()
            assert len(records) == 2
        finally:
            options_xlsx.XLSX_PATH = original_path

    def test_round_trip_save_load(self, xlsx_dir):
        filepath = xlsx_dir / "opt_roundtrip.xlsx"
        from app.data import options_xlsx
        original_path = options_xlsx.XLSX_PATH
        options_xlsx.XLSX_PATH = filepath
        try:
            original = [
                OptionsRecord(year=2022, amount=3500.0),
                OptionsRecord(year=2023, amount=4200.0),
            ]
            options_xlsx.save_options(original)
            loaded = options_xlsx.load_options()
            assert len(loaded) == 2
            assert loaded[0].amount == pytest.approx(3500.0)
        finally:
            options_xlsx.XLSX_PATH = original_path
